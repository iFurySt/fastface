from __future__ import annotations

import argparse
import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import cv2
import numpy as np
from PIL import Image, ImageFile

from fastface.paths import expand_path
from fastface.pipeline.detectors import DetectedFace, build_detector

ImageFile.LOAD_TRUNCATED_IMAGES = True
GENDER_NAMES = {0: "female", 1: "male"}


def softmax_numpy(logits: np.ndarray, axis: int = 1) -> np.ndarray:
    logits = logits - np.max(logits, axis=axis, keepdims=True)
    exp = np.exp(logits)
    return exp / np.sum(exp, axis=axis, keepdims=True).clip(min=1e-8)


@dataclass(frozen=True)
class AvatarItem:
    sample_id: str
    image_path: Path
    relative_path: str
    split: str
    sha256: str
    width: int
    height: int
    source_dataset_index: str
    source_archive_index: str


def load_avatar_items(dataset_root: Path, max_images: int | None = None, split: str | None = None) -> list[AvatarItem]:
    metadata_path = dataset_root / "metadata.csv"
    items: list[AvatarItem] = []
    with metadata_path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            if split and row.get("split") != split:
                continue
            relative_path = row["file_name"]
            items.append(
                AvatarItem(
                    sample_id=Path(relative_path).stem,
                    image_path=dataset_root / relative_path,
                    relative_path=relative_path,
                    split=row.get("split", ""),
                    sha256=row.get("sha256", ""),
                    width=int(row.get("width") or 0),
                    height=int(row.get("height") or 0),
                    source_dataset_index=row.get("source_dataset_index", ""),
                    source_archive_index=row.get("source_archive_index", ""),
                )
            )
            if max_images is not None and len(items) >= max_images:
                break
    return items


def preprocess_public_fairface(face_bgr: np.ndarray, input_size: int) -> np.ndarray:
    image = Image.fromarray(cv2.cvtColor(face_bgr, cv2.COLOR_BGR2RGB))
    array = np.asarray(image.resize((input_size, input_size), Image.Resampling.BILINEAR), dtype=np.float32) / 255.0
    mean = np.asarray([0.485, 0.456, 0.406], dtype=np.float32)
    std = np.asarray([0.229, 0.224, 0.225], dtype=np.float32)
    array = (array - mean) / std
    return np.transpose(array, (2, 0, 1)).astype(np.float32)


class PublicFairFaceOnnxPredictor:
    def __init__(self, model_path: Path, input_size: int = 224) -> None:
        try:
            import onnxruntime as ort
        except ImportError as exc:
            raise RuntimeError("onnxruntime is required for public FairFace ONNX prediction") from exc

        self.model_path = expand_path(model_path)
        self.input_size = input_size
        self.session = ort.InferenceSession(str(self.model_path), providers=["CPUExecutionProvider"])
        self.input_name = self.session.get_inputs()[0].name
        self.output_names = [output.name for output in self.session.get_outputs()]

    def predict_batch(self, face_bgrs: list[np.ndarray]) -> list[dict[str, Any]]:
        if not face_bgrs:
            return []
        batch = np.stack([preprocess_public_fairface(face, self.input_size) for face in face_bgrs], axis=0)
        outputs = self.session.run(self.output_names, {self.input_name: batch})
        if len(outputs) < 3:
            raise RuntimeError(f"public FairFace ONNX expected race/gender/age outputs, got {len(outputs)}")
        gender_probs_public = softmax_numpy(np.asarray(outputs[1]), axis=1)
        age_probs = softmax_numpy(np.asarray(outputs[2]), axis=1)
        predictions: list[dict[str, Any]] = []
        for gender_probs, age_prob in zip(gender_probs_public, age_probs):
            # Public FairFace gender order is ["Male", "Female"]; FastFace uses 0=female, 1=male.
            male_prob = float(gender_probs[0])
            female_prob = float(gender_probs[1])
            gender = 1 if male_prob >= female_prob else 0
            predictions.append(
                {
                    "gender": gender,
                    "gender_name": GENDER_NAMES[gender],
                    "female_prob": female_prob,
                    "male_prob": male_prob,
                    "gender_confidence": max(female_prob, male_prob),
                    "age_group": int(np.argmax(age_prob)),
                }
            )
        return predictions


def choose_face(faces: list[DetectedFace]) -> DetectedFace | None:
    if not faces:
        return None
    return max(faces, key=lambda face: (face.bbox[2] - face.bbox[0]) * (face.bbox[3] - face.bbox[1]) * face.score)


def base_record(item: AvatarItem) -> dict[str, Any]:
    return {
        "sample_id": item.sample_id,
        "image_path": str(item.image_path),
        "relative_path": item.relative_path,
        "split": item.split,
        "sha256": item.sha256,
        "width": item.width,
        "height": item.height,
        "source_dataset_index": item.source_dataset_index,
        "source_archive_index": item.source_archive_index,
    }


def flatten_record(record: dict[str, Any]) -> dict[str, Any]:
    flat = dict(record)
    for key in ["fastface", "public_fairface", "face"]:
        value = flat.pop(key, None)
        if isinstance(value, dict):
            for nested_key, nested_value in value.items():
                if isinstance(nested_value, (list, dict)):
                    flat[f"{key}_{nested_key}"] = json.dumps(nested_value, sort_keys=True)
                else:
                    flat[f"{key}_{nested_key}"] = nested_value
    return flat


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, sort_keys=True) + "\n")


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    flattened = [flatten_record(row) for row in rows]
    fieldnames = sorted({key for row in flattened for key in row})
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(flattened)


def make_thumbnail(source_image: Path, output_path: Path, size: int) -> None:
    try:
        image = Image.open(source_image).convert("RGB")
        image.thumbnail((size, size), Image.Resampling.LANCZOS)
    except Exception:
        image = Image.new("RGB", (size, size), "#dddddd")
    canvas = Image.new("RGB", (size, size), "white")
    canvas.paste(image, ((size - image.width) // 2, (size - image.height) // 2))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path, quality=88)


def write_review_workbook(path: Path, rows: list[dict[str, Any]], thumbnail_size: int) -> None:
    if not rows:
        return
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to write --review-xlsx") from exc

    image_dir = path.parent / "review_images"
    records = [flatten_record(row) for row in rows]
    for index, row in enumerate(records, start=1):
        thumbnail_path = image_dir / f"{index:06d}_{row['sample_id']}.jpg"
        make_thumbnail(Path(row["image_path"]), thumbnail_path, size=thumbnail_size)
        row["thumbnail"] = str(thumbnail_path.relative_to(path.parent))
        row["manual_gender"] = ""
        row["manual_note"] = ""

    preferred = [
        "thumbnail",
        "manual_gender",
        "manual_note",
        "review_reason",
        "sample_id",
        "relative_path",
        "split",
        "face_count",
        "fastface_gender_name",
        "fastface_gender_confidence",
        "public_fairface_gender_name",
        "public_fairface_gender_confidence",
        "fastface_age",
        "public_fairface_age_group",
        "detector_score",
    ]
    headers = preferred + sorted({key for row in records for key in row} - set(preferred))

    wb = Workbook()
    ws = wb.active
    ws.title = "review"
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    image_col = headers.index("thumbnail") + 1
    manual_col = headers.index("manual_gender") + 1
    for row_index, record in enumerate(records, start=2):
        ws.append([record.get(header, "") for header in headers])
        ws.row_dimensions[row_index].height = 140
        thumbnail_path = path.parent / record["thumbnail"]
        xl_image = XLImage(str(thumbnail_path))
        xl_image.width = 132
        xl_image.height = 132
        ws.add_image(xl_image, f"{get_column_letter(image_col)}{row_index}")
        ws.cell(row=row_index, column=image_col).value = None
        for col_index in range(1, len(headers) + 1):
            ws.cell(row=row_index, column=col_index).alignment = Alignment(vertical="top", wrap_text=True)
    validation = DataValidation(type="list", formula1='"female,male,unclear,not_face"', allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(f"{get_column_letter(manual_col)}2:{get_column_letter(manual_col)}{len(records) + 1}")
    widths = {"thumbnail": 22, "manual_gender": 16, "manual_note": 32, "review_reason": 28, "image_path": 72}
    for index, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(index)].width = widths.get(header, 18)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Label an avatar dataset by agreement between FastFace and a public baseline.")
    parser.add_argument("--dataset-root", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--fastface-model", type=Path, required=True)
    parser.add_argument("--public-fairface-onnx", type=Path, required=True)
    parser.add_argument("--detector-onnx", type=Path, required=True)
    parser.add_argument("--detector-input-size", type=int, default=1280)
    parser.add_argument("--detector-conf", type=float, default=0.65)
    parser.add_argument("--detector-nms", type=float, default=0.3)
    parser.add_argument("--detector-pre-nms-topk", type=int, default=1000)
    parser.add_argument("--split", choices=["train", "validation"], help="Optional split filter.")
    parser.add_argument("--max-images", type=int)
    parser.add_argument("--batch-size", type=int, default=64)
    parser.add_argument("--min-model-confidence", type=float, default=0.80)
    parser.add_argument("--review-xlsx", action="store_true")
    parser.add_argument("--thumbnail-size", type=int, default=180)
    args = parser.parse_args()

    from fastface.pipeline.runtime import FastFaceOnnxPredictor, align_face

    items = load_avatar_items(expand_path(args.dataset_root), max_images=args.max_images, split=args.split)
    detector = build_detector(
        backend="owned-retinaface-onnx",
        model_name=None,
        confidence_threshold=args.detector_conf,
        nms_threshold=args.detector_nms,
        input_size=args.detector_input_size,
        providers=["CPUExecutionProvider"],
        owned_onnx_path=str(expand_path(args.detector_onnx)),
        resize_mode="max-side",
        pre_nms_topk=args.detector_pre_nms_topk,
    )
    fastface = FastFaceOnnxPredictor(model_path=expand_path(args.fastface_model))
    public = PublicFairFaceOnnxPredictor(model_path=expand_path(args.public_fairface_onnx))

    accepted: list[dict[str, Any]] = []
    review: list[dict[str, Any]] = []
    no_face: list[dict[str, Any]] = []
    batch_faces: list[np.ndarray] = []
    batch_context: list[tuple[AvatarItem, dict[str, Any], np.ndarray]] = []

    def flush_batch() -> None:
        if not batch_faces:
            return
        fastface_predictions = [fastface.predict(face) for face in batch_faces]
        public_predictions = public.predict_batch(batch_faces)
        for (item, record, _), fastface_prediction, public_prediction in zip(batch_context, fastface_predictions, public_predictions):
            fastface_dict = {
                "gender": fastface_prediction.gender,
                "gender_name": fastface_prediction.gender_name,
                "female_prob": fastface_prediction.female_prob,
                "male_prob": fastface_prediction.male_prob,
                "gender_confidence": fastface_prediction.gender_confidence,
                "age": fastface_prediction.age,
            }
            record["fastface"] = fastface_dict
            record["public_fairface"] = public_prediction
            same_gender = fastface_prediction.gender == int(public_prediction["gender"])
            confident = (
                fastface_prediction.gender_confidence >= args.min_model_confidence
                and float(public_prediction["gender_confidence"]) >= args.min_model_confidence
            )
            if same_gender and confident:
                record["label_status"] = "accepted"
                record["pseudo_gender"] = fastface_prediction.gender
                record["pseudo_gender_name"] = fastface_prediction.gender_name
                record["pseudo_age"] = fastface_prediction.age
                accepted.append(record)
            else:
                record["label_status"] = "review"
                reasons = []
                if not same_gender:
                    reasons.append("gender_disagreement")
                if not confident:
                    reasons.append("low_confidence")
                record["review_reason"] = "|".join(reasons)
                review.append(record)
        batch_faces.clear()
        batch_context.clear()

    for index, item in enumerate(items, start=1):
        record = base_record(item)
        image = cv2.imread(str(item.image_path), cv2.IMREAD_COLOR)
        if image is None:
            record["label_status"] = "review"
            record["review_reason"] = "invalid_image"
            review.append(record)
            continue
        faces = detector.detect(image, max_faces=0, selection_metric="max")
        record["face_count"] = len(faces)
        face = choose_face(faces)
        if face is None:
            record["label_status"] = "no_face"
            record["review_reason"] = "no_face"
            no_face.append(record)
            continue
        if len(faces) != 1:
            record["label_status"] = "review"
            record["review_reason"] = "multiple_faces"
            review.append(record)
            continue
        crop, crop_mode = align_face(image, face, output_size=fastface.input_size)
        record["face"] = {
            "bbox": [float(value) for value in face.bbox],
            "score": face.score,
            "landmarks": [[float(x), float(y)] for x, y in face.landmarks],
            "crop_mode": crop_mode,
        }
        record["detector_score"] = face.score
        batch_faces.append(crop)
        batch_context.append((item, record, crop))
        if len(batch_faces) >= args.batch_size:
            flush_batch()
        if index % 1000 == 0:
            print(
                json.dumps(
                    {
                        "processed": index,
                        "accepted": len(accepted),
                        "review": len(review),
                        "no_face": len(no_face),
                    },
                    sort_keys=True,
                ),
                flush=True,
            )
    flush_batch()

    output_dir = expand_path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    write_jsonl(output_dir / "accepted.jsonl", accepted)
    write_jsonl(output_dir / "review.jsonl", review)
    write_jsonl(output_dir / "no_face.jsonl", no_face)
    write_csv(output_dir / "accepted.csv", accepted)
    write_csv(output_dir / "review.csv", review)
    write_csv(output_dir / "no_face.csv", no_face)
    if args.review_xlsx:
        write_review_workbook(output_dir / "review.xlsx", review, thumbnail_size=args.thumbnail_size)
    summary = {
        "dataset_root": str(expand_path(args.dataset_root)),
        "items": len(items),
        "accepted": len(accepted),
        "review": len(review),
        "no_face": len(no_face),
        "accepted_rate": len(accepted) / max(len(items), 1),
        "review_rate": len(review) / max(len(items), 1),
        "no_face_rate": len(no_face) / max(len(items), 1),
        "min_model_confidence": args.min_model_confidence,
        "detector": detector.name,
        "detector_conf": args.detector_conf,
        "outputs": {
            "accepted_jsonl": str(output_dir / "accepted.jsonl"),
            "accepted_csv": str(output_dir / "accepted.csv"),
            "review_jsonl": str(output_dir / "review.jsonl"),
            "review_csv": str(output_dir / "review.csv"),
            "no_face_jsonl": str(output_dir / "no_face.jsonl"),
            "no_face_csv": str(output_dir / "no_face.csv"),
        },
    }
    (output_dir / "summary.json").write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps(summary, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
