from __future__ import annotations

import argparse
import csv
import json
import zipfile
from collections import OrderedDict
from pathlib import Path
from typing import Any

from PIL import Image as PILImage
from PIL import ImageFile

from fastface.data.manifest_dataset import crop_face
from fastface.paths import expand_path

ImageFile.LOAD_TRUNCATED_IMAGES = True


def safe_image_name(index: int, sample_id: str) -> str:
    safe_id = sample_id.replace(":", "_").replace("/", "_")
    return f"{index:04d}_{safe_id}.jpg"


def load_manifest_items(paths: list[Path]) -> dict[str, dict[str, Any]]:
    items: dict[str, dict[str, Any]] = {}
    for path in [expand_path(path) for path in paths]:
        with path.open(encoding="utf-8") as handle:
            for line in handle:
                if not line.strip():
                    continue
                row = json.loads(line)
                items[str(row.get("sample_id"))] = row
    return items


def load_focused_rows(focused_files: list[tuple[str, Path]]) -> tuple[list[dict[str, str]], dict[str, int]]:
    rows: OrderedDict[str, dict[str, str]] = OrderedDict()
    source_counts: dict[str, int] = {}
    for bucket, path in focused_files:
        with path.open(newline="", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            count = 0
            for row in reader:
                count += 1
                sample_id = row["sample_id"]
                if sample_id not in rows:
                    copied = dict(row)
                    copied["review_buckets"] = bucket
                    rows[sample_id] = copied
                    continue
                existing = rows[sample_id]
                buckets = existing["review_buckets"].split("|")
                if bucket not in buckets:
                    existing["review_buckets"] += f"|{bucket}"
            source_counts[bucket] = count
    return list(rows.values()), source_counts


def make_thumbnail(
    source_image: Path,
    manifest_item: dict[str, Any] | None,
    output_path: Path,
    crop_margin: float,
    size: int,
) -> None:
    image = PILImage.open(source_image).convert("RGB")
    if manifest_item is not None:
        image = crop_face(image, manifest_item, margin=crop_margin)
    image.thumbnail((size, size), PILImage.Resampling.LANCZOS)
    canvas = PILImage.new("RGB", (size, size), "white")
    canvas.paste(image, ((size - image.width) // 2, (size - image.height) // 2))
    canvas.save(output_path, quality=88)


def parse_prediction_field(value: str) -> dict[str, str]:
    result: dict[str, str] = {}
    for part in value.split("|"):
        if ":" not in part:
            continue
        key, item_value = part.split(":", 1)
        result[key] = item_value
    return result


def build_records(
    rows: list[dict[str, str]],
    manifest_items: dict[str, dict[str, Any]],
    output_dir: Path,
    crop_margin: float,
    thumbnail_size: int,
) -> list[dict[str, Any]]:
    images_dir = output_dir / "images"
    images_dir.mkdir(parents=True, exist_ok=True)
    records: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        sample_id = row["sample_id"]
        gender_predictions = parse_prediction_field(row["gender_predictions"])
        male_probs = parse_prediction_field(row["male_probs"])
        image_name = safe_image_name(index, sample_id)
        image_output = images_dir / image_name
        try:
            make_thumbnail(
                source_image=expand_path(row["image_path"]),
                manifest_item=manifest_items.get(sample_id),
                output_path=image_output,
                crop_margin=crop_margin,
                size=thumbnail_size,
            )
        except Exception:
            fallback = PILImage.new("RGB", (thumbnail_size, thumbnail_size), "#dddddd")
            fallback.save(image_output, quality=88)

        records.append(
            {
                "review_buckets": row["review_buckets"],
                "image": str(Path("images") / image_name),
                "manual_gender": "",
                "sample_id": sample_id,
                "dataset": row["dataset"],
                "label_gender_name": row["label_gender_name"],
                "label_age": float(row["label_age"]),
                "our_large_gender": gender_predictions.get("our_large128_imdb_distill", ""),
                "public_fairface_gender": gender_predictions.get("public_fairface_onnx", ""),
                "our_large_male_prob": float(male_probs.get("our_large128_imdb_distill", "nan")),
                "public_fairface_male_prob": float(male_probs.get("public_fairface_onnx", "nan")),
                "teacher_gender": gender_predictions.get("teacher_v2s_imdb", ""),
                "small112_gender": gender_predictions.get("our_small112_imdb_distill", ""),
                "gender_predictions": row["gender_predictions"],
                "male_probs": row["male_probs"],
                "correct_models": row["correct_models"],
                "wrong_models": row["wrong_models"],
                "num_wrong_models": int(row["num_wrong_models"]),
                "confidence_gap": float(row["confidence_gap"]),
            }
        )
    return records


def write_csv(path: Path, records: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(records[0]))
        writer.writeheader()
        writer.writerows(records)


def write_xlsx(path: Path, records: list[dict[str, Any]], source_counts: dict[str, int], thumbnail_size: int) -> None:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to write the manual review workbook") from exc

    output_dir = path.parent
    headers = list(records[0])
    wb = Workbook()
    ws = wb.active
    ws.title = "review"
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

    image_col = headers.index("image") + 1
    manual_col = headers.index("manual_gender") + 1
    for row_index, record in enumerate(records, start=2):
        ws.append([record[header] for header in headers])
        ws.row_dimensions[row_index].height = 140
        xl_image = XLImage(str(output_dir / record["image"]))
        xl_image.width = 132
        xl_image.height = 132
        ws.add_image(xl_image, f"{get_column_letter(image_col)}{row_index}")
        ws.cell(row=row_index, column=image_col).value = None
        for col_index in range(1, len(headers) + 1):
            ws.cell(row=row_index, column=col_index).alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "review_buckets": 28,
        "image": 22,
        "manual_gender": 16,
        "sample_id": 26,
        "dataset": 14,
        "label_gender_name": 16,
        "label_age": 10,
        "our_large_gender": 18,
        "public_fairface_gender": 20,
        "our_large_male_prob": 18,
        "public_fairface_male_prob": 22,
        "teacher_gender": 16,
        "small112_gender": 16,
        "gender_predictions": 72,
        "male_probs": 72,
        "correct_models": 34,
        "wrong_models": 42,
        "num_wrong_models": 16,
        "confidence_gap": 16,
    }
    for index, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(index)].width = widths.get(header, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.sheet_view.showGridLines = False
    validation = DataValidation(type="list", formula1='"female,male,unclear,unknown"', allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(f"{get_column_letter(manual_col)}2:{get_column_letter(manual_col)}{len(records) + 1}")

    summary = wb.create_sheet("summary")
    summary_rows = [
        ["field", "value"],
        ["unique_review_rows", len(records)],
        *[[f"{name}_source_rows", count] for name, count in sorted(source_counts.items())],
        ["manual_gender_allowed_values", "female, male, unclear, unknown"],
        ["image", "Embedded face-crop thumbnail; CSV image column points to images/*.jpg"],
    ]
    for row in summary_rows:
        summary.append(row)
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    summary.column_dimensions["A"].width = 38
    summary.column_dimensions["B"].width = 80
    summary.freeze_panes = "A2"

    wb.save(path)
    loaded = load_workbook(path, read_only=False)
    if "review" not in loaded.sheetnames or loaded["review"].max_row != len(records) + 1:
        raise RuntimeError("workbook verification failed")


def write_zip(path: Path, output_dir: Path, files: list[Path]) -> None:
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for file_path in files:
            archive.write(file_path, file_path.name)
        for image_path in sorted((output_dir / "images").glob("*.jpg")):
            archive.write(image_path, str(Path("images") / image_path.name))


def main() -> None:
    parser = argparse.ArgumentParser(description="Build an image-embedded workbook for manual gender review.")
    parser.add_argument("--comparison-dir", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--manifest", type=Path, action="append", required=True)
    parser.add_argument(
        "--focused-file",
        action="append",
        help="Focused CSV as bucket:path. Defaults to public_vs_our_large from comparison-dir.",
    )
    parser.add_argument("--crop-margin", type=float, default=0.2)
    parser.add_argument("--thumbnail-size", type=int, default=180)
    args = parser.parse_args()

    if args.focused_file:
        focused_files = []
        for value in args.focused_file:
            bucket, path = value.split(":", 1)
            focused_files.append((bucket, expand_path(path)))
    else:
        focused_files = [("public_vs_our_large", args.comparison_dir / "focused" / "public_vs_our_large.csv")]
    args.output_dir.mkdir(parents=True, exist_ok=True)
    manifest_items = load_manifest_items(args.manifest)
    focused_rows, source_counts = load_focused_rows(focused_files)
    records = build_records(
        rows=focused_rows,
        manifest_items=manifest_items,
        output_dir=args.output_dir,
        crop_margin=args.crop_margin,
        thumbnail_size=args.thumbnail_size,
    )
    if not records:
        raise ValueError("no focused rows selected")

    csv_path = args.output_dir / "manual_gender_review.csv"
    xlsx_path = args.output_dir / "manual_gender_review.xlsx"
    zip_path = args.output_dir / "manual_gender_review_package.zip"
    write_csv(csv_path, records)
    write_xlsx(xlsx_path, records, source_counts, args.thumbnail_size)
    write_zip(zip_path, args.output_dir, [xlsx_path, csv_path])

    print(
        json.dumps(
            {
                "output_dir": str(args.output_dir),
                "xlsx": str(xlsx_path),
                "csv": str(csv_path),
                "zip": str(zip_path),
                "unique_review_rows": len(records),
                "source_counts": source_counts,
                "image_count": len(list((args.output_dir / "images").glob("*.jpg"))),
            },
            indent=2,
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
